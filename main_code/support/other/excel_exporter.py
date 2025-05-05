from main_code.well_model.simplified_well.simplified_well import SimplifiedBHE
from openpyxl import Workbook, load_workbook, styles
from abc import ABC, abstractmethod
from datetime import date, datetime
from tkinter import filedialog
import tkinter as tk
import pandas as pd
import numpy as np
import os


class ExportDict(ABC):

    def __init__(self):

        self.export_dict = dict()
        self.init_export_dict()

    @abstractmethod
    def init_export_dict(self, *args):
        pass

    @abstractmethod
    def append_values_to_dict(self, analysis_result, *args):
        pass


def export_to_excel(

        excel_dir_path="",
        excel_name="result.xlsx",
        add_date_time=False,
        sheet_name='data'

):

    def new_func(func):

        def new_wrapper(*args, **kwargs):

            if not os.path.isdir(excel_dir_path):

                root = tk.Tk()
                root.withdraw()

                excel_dir = filedialog.askdirectory()

            else:

                excel_dir = excel_dir_path

            if add_date_time:

                today = date.today()
                now = datetime.now()
                today_str = today.strftime("%d %b")
                now_str = now.strftime("%H.%M")

                new_excel_name = "{}_{}_{}".format(today_str, now_str, excel_name)

            else:

                new_excel_name = excel_name

            if ".xlsx" not in new_excel_name:
                new_excel_name = new_excel_name + ".xlsx"

            if os.path.isdir(excel_dir):

                output_file = os.path.join(excel_dir, new_excel_name)

                if os.path.isfile(output_file):

                    writer = pd.ExcelWriter(output_file, engine="openpyxl", mode="a", if_sheet_exists="replace")

                else:

                    writer = pd.ExcelWriter(output_file, engine="openpyxl", mode="w")

                result_dict = func(*args, **kwargs)
                df = pd.DataFrame(result_dict)
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                writer.save()

        return new_wrapper

    return new_func

def write_excel_sheet(

        excel_path, sheet_name, data_frame: dict,
        first_row=2, first_column=2, overwrite="light",
        write_data=True

):

    if not os.path.isfile(str(excel_path)):
        wb = Workbook()

    else:
        wb = load_workbook(excel_path)

    if overwrite == "hard":
        if sheet_name in wb.sheetnames:
            wb.remove_sheet(wb[sheet_name])

    elif overwrite == "none":
        if sheet_name in wb.sheetnames:
            return

    if not sheet_name in wb.sheetnames:
        wb.create_sheet(sheet_name)

    sheet = wb[sheet_name]

    if write_data:
        __write_excel_data(sheet, data_frame, first_row, first_column)
    else:
        __write_excel_specification(sheet, data_frame, first_row, first_column)

    wb.save(excel_path)


def __write_excel_data(sheet, data_frame: dict, first_row, first_column):

    col = first_column
    for key in data_frame.keys():

        row = first_row
        sub_data_frame = data_frame[key]
        n_sub_element = len(sub_data_frame["unit"])

        if n_sub_element == 0:

            sheet.merge_cells(start_row=row, start_column=col, end_row=row + 1, end_column=col)
            n_sub_element = 1

        else:

            if n_sub_element > 3:
                # Add a space between data entry if n_sub_element > 3
                col += 1

            if n_sub_element > 1:
                sheet.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + n_sub_element - 1)

            for n in range(n_sub_element):

                row = first_row + 1
                cell = sheet.cell(row, col + n, value=sub_data_frame["unit"][n])
                cell.alignment = styles.Alignment(horizontal="center", vertical="center")
                cell.font = styles.Font(italic=True, size=10)

        cell = sheet.cell(first_row, col, value=key)
        cell.alignment = styles.Alignment(horizontal="center", vertical="center")
        cell.font = styles.Font(bold=True)

        for n in range(n_sub_element):

            row = first_row + 3
            data_list = (sub_data_frame["values"])[n]

            for data in data_list:

                sheet.cell(row, col, value=data)
                row += 1

            col += 1


def __write_excel_specification(sheet, data_frame: dict, first_row, first_column):

    max_lines = 30

    data_col = 0
    row_offset = 0

    for key in data_frame.keys():

        if len(data_frame[key].keys()) > 0:

            if row_offset + len(data_frame[key].keys()) > max_lines:

                row_offset = 0
                data_col += 1

            name_col = first_column + data_col * 4
            value_col = first_column + data_col * 4 + 1
            units_col = first_column + data_col * 4 + 2
            sub_data_frame = data_frame[key]

            # Title
            row = first_row + row_offset
            sheet.merge_cells(start_row=row, start_column=name_col, end_row=row, end_column=units_col)
            cell = sheet.cell(row, name_col, value=key)
            cell.alignment = styles.Alignment(horizontal="center", vertical="center")
            cell.font = styles.Font(italic=True)

            for sub_keys in sub_data_frame.keys():

                row_offset += 1
                row = first_row + row_offset

                sheet.cell(row, value_col, value=sub_data_frame[sub_keys]["value"])
                cell = sheet.cell(row, name_col, value=sub_keys)
                cell.font = styles.Font(bold=True)

                if sub_data_frame[sub_keys]["unit"] is None:

                    sheet.merge_cells(start_row=row, start_column=value_col, end_row=row, end_column=units_col)
                    cell = sheet.cell(row, value_col)
                    cell.font = styles.Font(italic=True)
                    cell.alignment = styles.Alignment(horizontal="center", vertical="center")

                else:

                    cell = sheet.cell(row, units_col, value=sub_data_frame[sub_keys]["unit"])
                    cell.font = styles.Font(italic=True)

            row_offset += 2


def export_profiles_to_excel(

        file_path, data_input,
        times_in_main_tab=None, reverse_time_position=False,
        sheet_name='Main Results'

):

    well = data_input["well"]
    time_list = data_input["time_list"]
    t_out_list = data_input["t_out_list"]
    w_out_list = data_input["w_out_list"]
    p_out_list = data_input["p_out_list"]
    profile_positions = data_input["profile_positions"]

    # Write Specification Data Sheet
    write_excel_sheet(

        excel_path=file_path, sheet_name='Calculation Settings',
        data_frame=well.calculation_setup_data, overwrite="hard",
        write_data=False

    )

    if times_in_main_tab is None:

        main_time_list = time_list
        main_t_out_list = t_out_list
        main_p_out_list = p_out_list
        main_w_out_list = w_out_list

    else:

        main_time_list = list()
        main_t_out_list = list()
        main_p_out_list = list()
        main_w_out_list = list()

        for time in times_in_main_tab:

            if time in time_list:

                i = time_list.index(time)
                main_time_list.append(time_list[i])
                main_t_out_list.append(t_out_list[i])
                main_p_out_list.append(p_out_list[i])
                main_w_out_list.append(w_out_list[i])

    # Write Main Data Sheet
    main_data = {

        'Time': {"unit": ["days"], "values": [main_time_list]},
        'T_out': {"unit": ["°C"], "values": [main_t_out_list]},
        'P_out': {"unit": ["MPa"], "values": [main_p_out_list]},
        'W_out': {"unit": ["kW"], "values": [main_w_out_list]}

    }
    write_excel_sheet(excel_path=file_path, sheet_name=sheet_name, data_frame=main_data, overwrite="hard")

    profile_data = {

        "time_list": time_list,
        "profile_positions": profile_positions,
        "profiles_names": [],
        "profiles_list": [],

    }

    derivative_data = {

        "time_list": time_list,
        "profile_positions": profile_positions[:-1],
        "profiles_names": [],
        "profiles_list": [],

    }

    for key in data_input.keys():

        if "profile_list" in key:

            profile_data["profiles_list"].append(data_input[key])
            profile_data["profiles_names"].append('{} Profiles'.format(key.replace("_profile_list", "")))

        elif "derivative_list" in key:

            derivative_data["profiles_list"].append(data_input[key])
            derivative_data["profiles_names"].append('{} Derivatives'.format(key.replace("_derivative_list", "")))

    __write_profiles(file_path, profile_data, reverse_time_position=reverse_time_position)
    __write_profiles(file_path, derivative_data, reverse_time_position=reverse_time_position)


def __write_profiles(file_path, profile_data, reverse_time_position=False):

    time_list = profile_data["time_list"]
    profile_positions = profile_data["profile_positions"]
    n_days = len(time_list)

    for i in range(len(profile_data["profiles_list"])):

        curr_profile_name = profile_data["profiles_names"][i]
        curr_profile_list = np.array(profile_data["profiles_list"][i])

        if reverse_time_position:

            data ={'Position': {"unit": ["m"], "values": [profile_positions]}}

            for i in range(n_days):

                data.update({

                    "{:.2f} Days".format(time_list[i]): {

                        "unit": ["Annulus", "Tubing"], "values": curr_profile_list[i, :, :]

                    }

                })

            write_excel_sheet(

                excel_path=file_path, sheet_name=curr_profile_name,
                data_frame=data, overwrite="hard"

            )

        else:

            ann_data = {

                'Time': {"unit": ["days"], "values": [time_list]},
                'Annulus': {"unit": profile_positions, "values": curr_profile_list[:, 0, :].T},

            }
            tub_data = {

                'Time': {"unit": ["days"], "values": [time_list]},
                'Tubing': {"unit": profile_positions, "values": curr_profile_list[:, 1, :].T}

            }


            write_excel_sheet(

                excel_path=file_path, sheet_name=curr_profile_name,
                data_frame=ann_data, overwrite="hard"

            )

            write_excel_sheet(

                excel_path=file_path, sheet_name=curr_profile_name,
                data_frame=tub_data, first_row = 6 + n_days

            )
